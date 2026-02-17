import streamlit as st
import pandas as pd
import datetime
from datetime import timedelta
import numpy as np
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

#Streamlit
st.set_page_config(page_title="Apps Control de Asistencia",page_icon="logo_g.png", layout="wide")

st.title("📊 Control de Asistencia - Global")
st.markdown("Sube el archivo `.dat` para generar el reporte automatizado.")

#Definicion de variables

limite_fin_entrada_manana = datetime.time(11, 0, 0)
limite_fin_salida_manana = datetime.time(14, 0, 0)
limite_fin_entrada_tarde = datetime.time(17, 0, 0)
limite_fin_salida_manana_sab = datetime.time(15,0,0)
lu_vi = [0,1,2,3,4]

hora_default_entrada_m = datetime.time(9, 0, 0)
hora_default_salida_m = datetime.time(13, 0, 0)
hora_default_entrada_t = datetime.time(15, 0, 0)
hora_default_salida_t = datetime.time(19, 0, 0)
hora_default_salida_m_sab = datetime.time(14, 0, 0)

meta_ent_m = datetime.time(9, 0, 0)
meta_sal_m = datetime.time(13, 0, 0)
meta_ent_t = datetime.time(15, 0, 0)
meta_sal_t = datetime.time(19, 0, 0)
meta_sal_m_sab = datetime.time(14, 0, 0)

umbral_ent_m = datetime.time(9, 1, 0)
umbral_ent_t = datetime.time(15, 1, 0)

#Definicion de estructuras de datos

columnas_esperadas = [
    "sucio_entrada_mañana", 
    "sucio_salida_mañana", 
    "sucio_entrada_tarde", 
    "sucio_salida_tarde"
]

cols_totales = [
    "sucio_total_horas_mañana", 
    "sucio_total_horas_tarde",
    "redondeo_total_horas_mañana", 
    "redondeo_total_horas_tarde"
]

columnas_visuales = [
    "sucio_total_horas_mañana_hms", 
    "sucio_total_horas_tarde_hms",
    "redondeo_total_horas_mañana_hms", 
    "redondeo_total_horas_tarde_hms"
]

valores = ["08:00:00", "05:00:00", "00:00:00"]

nombres_dias = {
    0: 'Lunes', 1: 'Martes', 2: 'Miércoles', 3: 'Jueves',
    4: 'Viernes', 5: 'Sábado', 6: 'Domingo'
}

cols_tiempo = ['entrada_mañana', 'salida_mañana',
               'entrada_tarde', 'salida_tarde',
               'redondeo_entrada_mañana', 'redondeo_salida_mañana',
               'redondeo_entrada_tarde', 'redondeo_salida_tarde'
]

cols_totales_texto = [
    "sucio_total_horas_mañana_hms", 
    "sucio_total_horas_tarde_hms",
    "redondeo_total_horas_mañana_hms", 
    "redondeo_total_horas_tarde_hms", 
    "horas_requeridas"
]

columnas_finales = [
    "registros_bruto",     
    "nombre_dia",    
    "entrada_mañana",
    "salida_mañana",
    "entrada_tarde",
    "salida_tarde",
    "redondeo_entrada_mañana",
    "redondeo_salida_mañana",
    "redondeo_entrada_tarde",
    "redondeo_salida_tarde",
    "sucio_total_horas_mañana_hms",
    "sucio_total_horas_tarde_hms",
    "redondeo_total_horas_mañana_hms",
    "redondeo_total_horas_tarde_hms",
    "horas_requeridas"
]

cols_hora = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']

columnas = ["id_usuario", "fecha_hora", "col1", "col2", "col3", "col4"]

#Definicion de funciones


def clasificar_marcacion(fila):

    hora = fila["fecha_hora"].time()
    dia_semana = fila["fecha_hora"].weekday() 

    if dia_semana == 6:
        return None

    if hora < limite_fin_entrada_manana:
        return "sucio_entrada_mañana"
    
    elif (dia_semana in lu_vi) and (hora < limite_fin_salida_manana):
        return "sucio_salida_mañana"
    
    elif (dia_semana == 5) and (hora < limite_fin_salida_manana_sab):
        return "sucio_salida_mañana"
    
    elif hora < limite_fin_entrada_tarde:
        if dia_semana == 5: 
            return None 
        return "sucio_entrada_tarde"
    
    else: 
        if dia_semana == 5: 
            return None
        return "sucio_salida_tarde"
    
    
def rellenar_vacios(fila):
    fecha = fila.name 
    dia_semana = fecha.weekday() 
    
    
    nueva_fila = fila.copy()
    
    if dia_semana == 6:
        
        return nueva_fila 

    
    if pd.isna(nueva_fila["sucio_entrada_mañana"]):
        
        nueva_fila["sucio_entrada_mañana"] = pd.Timestamp.combine(fecha, hora_default_entrada_m)

    
    if pd.isna(nueva_fila["sucio_salida_mañana"]):
        if dia_semana == 5: 
            nueva_fila["sucio_salida_mañana"] = pd.Timestamp.combine(fecha, hora_default_salida_m_sab) #Augusto
        nueva_fila["sucio_salida_mañana"] = pd.Timestamp.combine(fecha, hora_default_salida_m)

    if dia_semana < 5: # 0,1,2,3,4
        if pd.isna(nueva_fila["sucio_entrada_tarde"]):
            nueva_fila["sucio_entrada_tarde"] = pd.Timestamp.combine(fecha, hora_default_entrada_t)
        
        if pd.isna(nueva_fila["sucio_salida_tarde"]):
            nueva_fila["sucio_salida_tarde"] = pd.Timestamp.combine(fecha, hora_default_salida_t)
            

    elif dia_semana == 5:       
        nueva_fila["sucio_entrada_tarde"] = pd.NaT
        nueva_fila["sucio_salida_tarde"] = pd.NaT

    return nueva_fila


def aplicar_redondeo(fila):
    fecha = fila.name #
    dia_semana = fecha.weekday()
    
    
    def hacer_ts(hora):
        return pd.Timestamp.combine(fecha, hora)

    
    res = {}


    val = fila["sucio_entrada_mañana"]
    if pd.isna(val):
        res["redondeo_entrada_mañana"] = pd.NaT
    elif val < hacer_ts(umbral_ent_m):
        res["redondeo_entrada_mañana"] = hacer_ts(meta_ent_m)
    else:
        res["redondeo_entrada_mañana"] = val

 
    val = fila["sucio_salida_mañana"]
    if pd.isna(val):
        res["redondeo_salida_mañana"] = pd.NaT
    elif (val >= hacer_ts(meta_sal_m)) and (dia_semana < 5):
        res["redondeo_salida_mañana"] = hacer_ts(meta_sal_m)
    elif (val >= hacer_ts(meta_sal_m_sab)) and (dia_semana == 5):
        res["redondeo_salida_mañana"] = hacer_ts(meta_sal_m_sab)
    else:
        res["redondeo_salida_mañana"] = val


    val = fila["sucio_entrada_tarde"]
    if pd.isna(val):
        res["redondeo_entrada_tarde"] = pd.NaT
    elif val < hacer_ts(umbral_ent_t):
        res["redondeo_entrada_tarde"] = hacer_ts(meta_ent_t)
    else:
        res["redondeo_entrada_tarde"] = val


    val = fila["sucio_salida_tarde"]
    if pd.isna(val):
        res["redondeo_salida_tarde"] = pd.NaT
    elif val >= hacer_ts(meta_sal_t):
        res["redondeo_salida_tarde"] = hacer_ts(meta_sal_t)
    else:
        res["redondeo_salida_tarde"] = val

    return pd.Series(res)


def timedelta_a_decimal(td):

    if pd.isna(td):
        return 0.0 

    return td.total_seconds() / 3600


def formatear_timedelta(td):
    if pd.isna(td):
        return "00:00:00" 
    
    total_segundos = int(td.total_seconds())
    
    horas = total_segundos // 3600
    minutos = (total_segundos % 3600) // 60
    segundos = total_segundos % 60
    
    
    return f"{horas:02}:{minutos:02}:{segundos:02}"

#streamlit
archivo = st.file_uploader("Carga el archivo .dat aquí", type=["dat", "txt"])

if archivo:
    df = pd.read_csv(archivo, sep = "\t", header = None, names=columnas)
    df = df[["id_usuario", "fecha_hora"]]
    df["id_usuario"] = df["id_usuario"].astype(str)

    with st.sidebar:
        st.header("Configuración del Reporte")
        nombre_personal = st.text_input("Nombre del Personal", value=None)
        
        usuarios_unicos = sorted(df["id_usuario"].unique(), key=lambda x: int(x) if x.isdigit() else 0)
        id_buscar = st.selectbox("Selecciona ID de Usuario", usuarios_unicos)
        
        col_f1, col_f2 = st.columns(2)
        fecha_inicio = col_f1.date_input("Fecha Inicio", value="today")
        fecha_final = col_f2.date_input("Fecha Final", value="today")
        fecha_inicio = pd.Timestamp(fecha_inicio)
        fecha_final = pd.Timestamp(fecha_final) + timedelta(days=1)
        rango_completo = pd.date_range(start=fecha_inicio, end=fecha_final, freq='D')

    if st.button("🚀 Generar Reporte"):
        with st.spinner("Procesando datos..."):

            df["fecha_hora"] = pd.to_datetime(df["fecha_hora"])
            df_id = df.loc[df["id_usuario"] == id_buscar].copy()

            df_filtrado = df_id[df_id["fecha_hora"].between(fecha_inicio, fecha_final)].copy()
            df_filtrado["fecha_hora"] = df_filtrado["fecha_hora"].dt.floor("min")
            df_filtrado["fecha"] = df_filtrado["fecha_hora"].dt.date
            df_filtrado["hora"] = df_filtrado["fecha_hora"].dt.time

            df_filtrado["categoria"] = df_filtrado.apply(clasificar_marcacion, axis=1).copy()
            df_clasificado = df_filtrado.dropna(subset=["categoria"]).copy()

            df_resultado_fase2 = df_clasificado.pivot_table(
                index="fecha",
                columns="categoria",
                values="fecha_hora",
                aggfunc="min"
            )

            df_resultado_fase2 = df_resultado_fase2.reindex(columns=columnas_esperadas)

            df_imputado = df_resultado_fase2.reindex(rango_completo)
            df_imputado.index.name = "fecha"

            df_final_fase2 = df_imputado.apply(rellenar_vacios, axis=1)

            df_inicial = df_imputado.copy()
            df_inicial.columns = ['entrada_mañana', 'salida_mañana', 'entrada_tarde', 'salida_tarde']

            cols_redondeo = df_final_fase2.apply(aplicar_redondeo, axis=1)
            df_final_fase3 = pd.concat([df_inicial, df_final_fase2, cols_redondeo], axis=1)

            df_final_fase3["sucio_total_horas_mañana"] = df_final_fase3["sucio_salida_mañana"] - df_final_fase3["sucio_entrada_mañana"]
            df_final_fase3["sucio_total_horas_tarde"] = df_final_fase3["sucio_salida_tarde"] - df_final_fase3["sucio_entrada_tarde"]

            df_final_fase3["redondeo_total_horas_mañana"] = df_final_fase3["redondeo_salida_mañana"] - df_final_fase3["redondeo_entrada_mañana"]
            df_final_fase3["redondeo_total_horas_tarde"] = df_final_fase3["redondeo_salida_tarde"] - df_final_fase3["redondeo_entrada_tarde"]

            for col in cols_totales:
                df_final_fase3[col] = df_final_fase3[col].apply(timedelta_a_decimal)

            df_calculado = df_final_fase3.copy()

            diferencia = df_final_fase3["sucio_salida_mañana"] - df_final_fase3["sucio_entrada_mañana"]
            df_final_fase3["sucio_total_horas_mañana_hms"] = diferencia.apply(formatear_timedelta)

            diferencia = df_final_fase3["sucio_salida_tarde"] - df_final_fase3["sucio_entrada_tarde"]
            df_final_fase3["sucio_total_horas_tarde_hms"] = diferencia.apply(formatear_timedelta)

            diferencia = df_final_fase3["redondeo_salida_mañana"] - df_final_fase3["redondeo_entrada_mañana"]
            df_final_fase3["redondeo_total_horas_mañana_hms"] = diferencia.apply(formatear_timedelta)

            diferencia = df_final_fase3["redondeo_salida_tarde"] - df_final_fase3["redondeo_entrada_tarde"]
            df_final_fase3["redondeo_total_horas_tarde_hms"] = diferencia.apply(formatear_timedelta)

            df_final_fase3["dia_semana"] = df_final_fase3.index.dayofweek

            df_final_fase3["nombre_dia"] = df_final_fase3["dia_semana"].map(nombres_dias)

            condiciones = [
                (df_final_fase3["dia_semana"] < 5),  
                (df_final_fase3["dia_semana"] == 5), 
                (df_final_fase3["dia_semana"] == 6)  
            ]

            df_final_fase3["horas_requeridas"] = np.select(condiciones, valores, default="00:00:00")

            #Parentesis
            df_filtrado["hora_texto"] = df_filtrado["fecha_hora"].dt.strftime('%H:%M:%S')

            columna_bruta = df_filtrado.groupby("fecha")["hora_texto"].apply(lambda x: ", ".join(x))

            columna_bruta.name = "registros_bruto"

            if "registros_bruto" in df_final_fase3.columns:
                df_final_fase3 = df_final_fase3.drop(columns=["registros_bruto"])

            df_final_fase3 = df_final_fase3.join(columna_bruta)

            df_reporte = df_final_fase3[columnas_finales].copy()
            df_reporte = df_reporte.iloc[:-1,:].copy()

            for col in cols_tiempo:
                if col in df_reporte.columns:

                    df_reporte[col] = pd.to_datetime(df_reporte[col])
                    df_reporte[col] = df_reporte[col].dt.strftime('%H:%M')

            for col in cols_totales_texto:
                if col in df_reporte.columns:

                    df_reporte[col] = df_reporte[col].astype(str).str[:5]

                    df_reporte[col] = df_reporte[col].replace("nan", "")
                    df_reporte[col] = df_reporte[col].replace("None", "")


            # Estilos
            output= BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_reporte.to_excel(writer, sheet_name="Asistencia_Detallada")

            #nombre_archivo = f"{nombre_personal}_{fecha_final}_reporte_asistencia.xlsx"
            #df_reporte.to_excel(nombre_archivo, sheet_name="Asistencia_Detallada")
            output.seek(0)
            #Formateo
            wb = load_workbook(output)
            ws = wb.active 

            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))
            weekend_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

            fill_azul = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")    # A, B, C, P
            fill_verde = PatternFill(start_color="375623", end_color="375623", fill_type="solid")   # D, E, F, G
            fill_naranja = PatternFill(start_color="843C0C", end_color="843C0C", fill_type="solid") # H, I, J, K
            fill_gris = PatternFill(start_color="3B3838", end_color="3B3838", fill_type="solid")    # L, M
            fill_morado = PatternFill(start_color="3F3F76", end_color="3F3F76", fill_type="solid")  # N, O

            for cell in ws[1]:
                letra = cell.column_letter
                cell.font = header_font
                cell.alignment = header_align
                cell.border = thin_border
                
                if letra in ['A', 'B', 'C', 'P']:
                    cell.fill = fill_azul
                elif letra in ['D', 'E', 'F', 'G']:
                    cell.fill = fill_verde
                elif letra in ['H', 'I', 'J', 'K']:
                    cell.fill = fill_naranja
                elif letra in ['L', 'M']:
                    cell.fill = fill_gris
                elif letra in ['N', 'O']:
                    cell.fill = fill_morado

            for col in ws.columns:
                letra_columna = col[0].column_letter
                ws.column_dimensions[letra_columna].width = 18

                for cell in col:
                    if cell.row == 1:
                        continue
                        
                    cell.border = thin_border
                    
                    if letra_columna == 'A':
                        cell.number_format = 'DD/MM/YYYY' # Cambia "31/01/2026 00:00:00" a "31/01/2026"
                        cell.alignment = Alignment(horizontal="center")
                    
                    if letra_columna in cols_hora:
                        cell.number_format = 'HH:MM:SS'
                        cell.alignment = Alignment(horizontal="center")

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                celda_dia = row[2] # Columna C
                
                if celda_dia.value == "Sábado" or celda_dia.value == "Domingo":
                    for cell in row[0:3]: # Solo Columnas A, B y C
                        cell.fill = weekend_fill
            
            final_excell = BytesIO()
            wb.save(final_excell)

            st.success(f"¡Reporte procesado con éxito! Nombre: {nombre_personal} ID: {id_buscar}")
            st.download_button(
                label="📥 Descargar Reporte en Excel",
                data=final_excell.getvalue(),
                file_name=f"{nombre_personal}_{fecha_final}_asistencia.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )